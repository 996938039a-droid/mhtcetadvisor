"""
MHT-CET Export Utilities

Handles exporting preference lists to:
1. Excel (.xlsx) - with formatting and multiple sheets
2. PDF - print-ready report for counseling sessions
"""

import io
from typing import List, Dict, Optional
from datetime import datetime
import pandas as pd

# Excel libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF libraries
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .recommendation_engine import (
    RecommendationResult, 
    PreferenceListItem,
    StudentProfile
)
from .probability_engine import AdmissionChance


class ExcelExporter:
    """
    Exports preference lists to formatted Excel files
    """
    
    # Color scheme
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    DREAM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    TARGET_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    SAFE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(size=10)
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.workbook = None
    
    def export(
        self,
        result: RecommendationResult,
        filepath: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export recommendation result to Excel file
        
        Args:
            result: RecommendationResult object
            filepath: Optional file path to save (returns BytesIO if None)
        
        Returns:
            BytesIO buffer containing Excel file
        """
        self.workbook = Workbook()
        
        # Create sheets
        self._create_summary_sheet(result)
        self._create_preference_list_sheet(result)
        self._create_analysis_sheet(result)
        
        # Remove default sheet if it exists
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        
        # Save to buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        if filepath:
            with open(filepath, 'wb') as f:
                f.write(buffer.getvalue())
        
        return buffer
    
    def _create_summary_sheet(self, result: RecommendationResult):
        """Create summary sheet with student profile and overview"""
        ws = self.workbook.create_sheet("Summary", 0)
        
        profile = result.student_profile
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "MHT-CET College Preference Advisor - Summary Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Generated timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, size=9)
        
        # Student Profile Section
        ws['A4'] = "STUDENT PROFILE"
        ws['A4'].font = Font(bold=True, size=12)
        
        profile_data = [
            ("Percentile", f"{profile.percentile:.2f}"),
            ("Category", profile.category),
            ("Gender", profile.gender.capitalize()),
            ("Quota", profile.quota),
            ("Home University", profile.home_university or "N/A"),
            ("TFWS Eligible", "Yes" if profile.has_tfws else "No"),
            ("PWD", "Yes" if profile.has_pwd else "No"),
            ("Defence Quota", "Yes" if profile.has_defence else "No"),
        ]
        
        row = 5
        for label, value in profile_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Recommendation Summary
        ws[f'A{row + 1}'] = "RECOMMENDATION SUMMARY"
        ws[f'A{row + 1}'].font = Font(bold=True, size=12)
        
        summary_data = [
            ("Total Options", result.summary['total']),
            ("Dream Options", result.summary['dream']),
            ("Target Options", result.summary['target']),
            ("Safe Options", result.summary['safe']),
            ("Government Colleges", result.summary.get('government', 0)),
            ("Cutoff Adjustment", f"{result.cutoff_adjustment:+.1f}%"),
        ]
        
        row += 2
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Strategy Notes
        if result.strategy_notes:
            row += 1
            ws[f'A{row}'] = "STRATEGY NOTES"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for note in result.strategy_notes:
                ws[f'A{row}'] = note
                row += 1
        
        # Warnings
        if result.warnings:
            row += 1
            ws[f'A{row}'] = "WARNINGS"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
            row += 1
            
            for warning in result.warnings:
                ws[f'A{row}'] = warning
                ws[f'A{row}'].font = Font(color="FF0000")
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_preference_list_sheet(self, result: RecommendationResult):
        """Create main preference list sheet"""
        ws = self.workbook.create_sheet("Preference List", 1)
        
        # Headers
        headers = [
            "Rank", "College Code", "College Name", "Branch Code", 
            "Branch Name", "Category", "Probability", "Chance", 
            "Trend", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.THIN_BORDER
        
        # Data rows
        for row_idx, item in enumerate(result.preference_list, 2):
            row_data = [
                item.rank,
                item.college_code,
                item.college_name,
                item.branch_code,
                item.branch_name,
                item.category_code,
                f"{item.probability:.1%}",
                item.chance_category.value,
                item.trend,
                "; ".join(item.notes[:2]) if item.notes else ""
            ]
            
            # Determine row color based on chance category
            if item.chance_category in [AdmissionChance.DREAM, AdmissionChance.REACH]:
                fill = self.DREAM_FILL
            elif item.chance_category == AdmissionChance.TARGET:
                fill = self.TARGET_FILL
            else:
                fill = self.SAFE_FILL
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = fill
                cell.font = self.NORMAL_FONT
                cell.border = self.THIN_BORDER
                
                if col == 1:  # Rank column
                    cell.alignment = Alignment(horizontal='center')
                elif col == 7:  # Probability
                    cell.alignment = Alignment(horizontal='right')
        
        # Adjust column widths
        widths = [6, 12, 40, 12, 35, 12, 12, 10, 10, 40]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_analysis_sheet(self, result: RecommendationResult):
        """Create analysis sheet with trends and statistics"""
        ws = self.workbook.create_sheet("Analysis", 2)
        
        ws['A1'] = "CUTOFF TREND ANALYSIS"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Group by trend
        trend_counts = {}
        for item in result.preference_list:
            trend = item.trend
            trend_counts[trend] = trend_counts.get(trend, 0) + 1
        
        row = 3
        ws['A3'] = "Trend"
        ws['B3'] = "Count"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        row = 4
        for trend, count in trend_counts.items():
            ws[f'A{row}'] = trend.capitalize()
            ws[f'B{row}'] = count
            row += 1
        
        # Probability distribution
        row += 2
        ws[f'A{row}'] = "PROBABILITY DISTRIBUTION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        prob_ranges = {
            "90-100%": 0,
            "70-90%": 0,
            "50-70%": 0,
            "30-50%": 0,
            "10-30%": 0,
            "0-10%": 0
        }
        
        for item in result.preference_list:
            p = item.probability
            if p >= 0.9:
                prob_ranges["90-100%"] += 1
            elif p >= 0.7:
                prob_ranges["70-90%"] += 1
            elif p >= 0.5:
                prob_ranges["50-70%"] += 1
            elif p >= 0.3:
                prob_ranges["30-50%"] += 1
            elif p >= 0.1:
                prob_ranges["10-30%"] += 1
            else:
                prob_ranges["0-10%"] += 1
        
        row += 1
        ws[f'A{row}'] = "Range"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        for range_name, count in prob_ranges.items():
            ws[f'A{row}'] = range_name
            ws[f'B{row}'] = count
            row += 1


class PDFExporter:
    """
    Exports preference lists to formatted PDF reports
    """
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='Title',
            parent=self.styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20
        ))
        
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=15
        ))
        
        self.styles.add(ParagraphStyle(
            name='SmallText',
            parent=self.styles['Normal'],
            fontSize=8
        ))
    
    def export(
        self,
        result: RecommendationResult,
        filepath: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export recommendation result to PDF file
        
        Args:
            result: RecommendationResult object
            filepath: Optional file path to save
        
        Returns:
            BytesIO buffer containing PDF file
        """
        buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph(
            "MHT-CET College Preference Advisor",
            self.styles['Title']
        ))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['Normal']
        ))
        elements.append(Spacer(1, 20))
        
        # Student Profile
        elements.append(Paragraph("Student Profile", self.styles['SectionHeader']))
        elements.extend(self._create_profile_table(result.student_profile))
        elements.append(Spacer(1, 15))
        
        # Summary
        elements.append(Paragraph("Recommendation Summary", self.styles['SectionHeader']))
        elements.extend(self._create_summary_table(result))
        elements.append(Spacer(1, 15))
        
        # Strategy Notes
        if result.strategy_notes:
            elements.append(Paragraph("Strategy Notes", self.styles['SectionHeader']))
            for note in result.strategy_notes:
                elements.append(Paragraph(f"• {note}", self.styles['Normal']))
            elements.append(Spacer(1, 10))
        
        # Warnings
        if result.warnings:
            elements.append(Paragraph("Important Warnings", self.styles['SectionHeader']))
            for warning in result.warnings:
                elements.append(Paragraph(
                    f"⚠️ {warning}", 
                    ParagraphStyle('Warning', parent=self.styles['Normal'], textColor=colors.red)
                ))
            elements.append(Spacer(1, 10))
        
        # Page break before preference list
        elements.append(PageBreak())
        
        # Preference List
        elements.append(Paragraph("Preference List", self.styles['SectionHeader']))
        elements.extend(self._create_preference_table(result.preference_list))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        if filepath:
            with open(filepath, 'wb') as f:
                f.write(buffer.getvalue())
        
        return buffer
    
    def _create_profile_table(self, profile: StudentProfile) -> list:
        """Create student profile table"""
        data = [
            ["Percentile", f"{profile.percentile:.2f}"],
            ["Category", profile.category],
            ["Gender", profile.gender.capitalize()],
            ["Quota", profile.quota],
            ["Home University", profile.home_university or "N/A"],
            ["Special Quotas", self._get_special_quotas_str(profile)],
        ]
        
        table = Table(data, colWidths=[3*cm, 10*cm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ]))
        
        return [table]
    
    def _get_special_quotas_str(self, profile: StudentProfile) -> str:
        """Get string representation of special quotas"""
        quotas = []
        if profile.has_tfws:
            quotas.append("TFWS")
        if profile.has_pwd:
            quotas.append("PWD")
        if profile.has_defence:
            quotas.append("Defence")
        if profile.is_orphan:
            quotas.append("Orphan")
        return ", ".join(quotas) if quotas else "None"
    
    def _create_summary_table(self, result: RecommendationResult) -> list:
        """Create summary statistics table"""
        data = [
            ["Total Options", str(result.summary['total'])],
            ["Dream Options", str(result.summary['dream'])],
            ["Target Options", str(result.summary['target'])],
            ["Safe Options", str(result.summary['safe'])],
            ["Government Colleges", str(result.summary.get('government', 0))],
            ["Cutoff Adjustment", f"{result.cutoff_adjustment:+.1f}%"],
        ]
        
        table = Table(data, colWidths=[4*cm, 3*cm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        return [table]
    
    def _create_preference_table(self, preference_list: List[PreferenceListItem]) -> list:
        """Create preference list table"""
        # Header
        data = [["#", "College", "Branch", "Category", "Prob.", "Chance"]]
        
        # Define colors for chance categories
        chance_colors = {
            AdmissionChance.DREAM: colors.Color(1, 0.95, 0.8),
            AdmissionChance.REACH: colors.Color(1, 0.9, 0.7),
            AdmissionChance.TARGET: colors.Color(0.9, 0.95, 0.85),
            AdmissionChance.SAFE: colors.Color(0.85, 0.92, 0.98),
            AdmissionChance.ASSURED: colors.Color(0.8, 0.95, 0.8),
        }
        
        row_colors = []
        
        for item in preference_list:
            # Truncate long names
            college_short = item.college_name[:35] + "..." if len(item.college_name) > 38 else item.college_name
            branch_short = item.branch_name[:25] + "..." if len(item.branch_name) > 28 else item.branch_name
            
            data.append([
                str(item.rank),
                college_short,
                branch_short,
                item.category_code,
                f"{item.probability:.0%}",
                item.chance_category.value
            ])
            
            row_colors.append(chance_colors.get(item.chance_category, colors.white))
        
        table = Table(data, colWidths=[0.8*cm, 6*cm, 4.5*cm, 2*cm, 1.2*cm, 1.5*cm])
        
        # Base style
        style_commands = [
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.12, 0.31, 0.47)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        ]
        
        # Add row-specific colors based on chance category
        for i, color in enumerate(row_colors, 1):
            style_commands.append(('BACKGROUND', (0, i), (-1, i), color))
        
        table.setStyle(TableStyle(style_commands))
        
        return [table]


def export_to_excel(result: RecommendationResult, filepath: Optional[str] = None) -> io.BytesIO:
    """Convenience function for Excel export"""
    exporter = ExcelExporter()
    return exporter.export(result, filepath)


def export_to_pdf(result: RecommendationResult, filepath: Optional[str] = None) -> io.BytesIO:
    """Convenience function for PDF export"""
    exporter = PDFExporter()
    return exporter.export(result, filepath)


if __name__ == "__main__":
    # Test exporters
    from .data_loader import create_sample_data
    from .probability_engine import ProbabilityEngine
    from .recommendation_engine import RecommendationEngine, StudentProfile, PriorityType
    
    # Create sample data
    sample_df = create_sample_data()
    prob_engine = ProbabilityEngine(sample_df)
    rec_engine = RecommendationEngine(prob_engine)
    
    # Create sample profile
    profile = StudentProfile(
        percentile=96.5,
        category="OBC",
        gender="male",
        quota="MH",
        home_university="Sant Gadge Baba Amravati University",
        has_tfws=True,
        priority_type=PriorityType.BALANCED
    )
    
    # Generate recommendations
    result = rec_engine.generate_recommendations(profile)
    
    # Test Excel export
    excel_buffer = export_to_excel(result, "test_export.xlsx")
    print(f"Excel exported: {len(excel_buffer.getvalue())} bytes")
    
    # Test PDF export
    pdf_buffer = export_to_pdf(result, "test_export.pdf")
    print(f"PDF exported: {len(pdf_buffer.getvalue())} bytes")
